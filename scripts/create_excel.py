import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Planning Seminaire"

headers = ["Jour", "Date", "Heure", "Séquence / Thème", "Intervenant", "Durée (h)", "Contact"]
ws.append(headers)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

data = [
    ["Jour 1", "10/03/2026", "09:00", "Accueil & Introduction au séminaire", "Directeur", 1, "dir@entreprise.com"],
    ["Jour 1", "10/03/2026", "10:00", "Stratégie Annuelle", "Jean Dupont", 2.5, "jean.d@entreprise.com"],
    ["Jour 1", "10/03/2026", "14:00", "Atelier Innovation", "Marie Curie", 3, "marie.c@entreprise.com"],
    ["Jour 2", "11/03/2026", "09:00", "Transformation Numérique", "Alan Turing", 3.5, "alan.t@entreprise.com"],
    ["Jour 2", "11/03/2026", "14:00", "Intelligence Artificielle & Outils", "Ada Lovelace", 3, "ada.l@entreprise.com"],
    ["Jour 3", "12/03/2026", "09:00", "Management Agile", "Steve Jobs", 4, "steve.j@entreprise.com"],
    ["Jour 3", "12/03/2026", "14:00", "Cohésion d'équipe", "Coach Sportif", 3.5, "coach@entreprise.com"],
    ["Jour 4", "13/03/2026", "09:00", "Définition des objectifs Q2", "Directeur", 3, "dir@entreprise.com"],
    ["Jour 4", "13/03/2026", "14:00", "Clôture et remise des supports", "RH", 2, "rh@entreprise.com"]
]

for row in data:
    ws.append(row)

for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 25

wb.save(r"C:\Users\Administrator\Documents\twisterlab\Seminaire_Formation.xlsx")
print("Excel créé!")
